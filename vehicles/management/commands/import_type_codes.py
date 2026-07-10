import openpyxl
from django.core.management.base import BaseCommand, CommandError
from django.db import transaction

from vehicles.models import VehicleTypeCode

DEFAULT_SHEET = 'Smarka'
HEADER_ROWS = 2  # 1. satır dönem başlığı, 2. satır kolon adları


class Command(BaseCommand):
    help = 'Resmi marka/tip kodu xlsx dosyasını (Marka Kodu, Tip Kodu, Marka Adı, Tip Adı) VehicleTypeCode tablosuna aktarır.'

    def add_arguments(self, parser):
        parser.add_argument('xlsx_path', type=str)
        parser.add_argument('--sheet', type=str, default=DEFAULT_SHEET)

    def handle(self, *args, **options):
        path = options['xlsx_path']
        sheet_name = options['sheet']
        try:
            wb = openpyxl.load_workbook(path, read_only=True, data_only=True)
        except FileNotFoundError:
            raise CommandError(f"Dosya bulunamadı: {path}")

        if sheet_name not in wb.sheetnames:
            raise CommandError(f"'{sheet_name}' sayfası bulunamadı. Mevcut sayfalar: {wb.sheetnames}")

        ws = wb[sheet_name]
        all_rows = ws.iter_rows(min_row=1, max_row=HEADER_ROWS, values_only=True)
        next(all_rows)  # dönem başlığı satırı
        kolon_basliklari = next(all_rows)  # ..., 2026, 2025, ..., 2012
        yil_kolonlari = [
            (i, int(baslik)) for i, baslik in enumerate(kolon_basliklari)
            if isinstance(baslik, (int, float))
        ]

        rows = ws.iter_rows(min_row=HEADER_ROWS + 1, values_only=True)

        to_create = []
        seen = set()
        skipped = 0
        for row in rows:
            marka_kodu, tip_kodu, marka_adi, tip_adi = row[0], row[1], row[2], row[3]
            if marka_kodu is None or tip_kodu is None or not marka_adi or not tip_adi:
                skipped += 1
                continue
            key = (int(marka_kodu), int(tip_kodu))
            if key in seen:
                skipped += 1
                continue
            seen.add(key)
            yillar = [yil for i, yil in yil_kolonlari if i < len(row) and row[i]]
            to_create.append(VehicleTypeCode(
                marka_kodu=key[0], tip_kodu=key[1],
                marka_adi=str(marka_adi).strip(), tip_adi=str(tip_adi).strip(),
                ilk_yil=min(yillar) if yillar else None,
                son_yil=max(yillar) if yillar else None,
            ))

        with transaction.atomic():
            VehicleTypeCode.objects.bulk_create(
                to_create, batch_size=1000,
                update_conflicts=True,
                unique_fields=['marka_kodu', 'tip_kodu'],
                update_fields=['marka_adi', 'tip_adi', 'ilk_yil', 'son_yil'],
            )

        self.stdout.write(self.style.SUCCESS(
            f"{len(to_create)} tip kodu içe aktarıldı/güncellendi ({skipped} satır atlandı)."
        ))
